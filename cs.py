from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = "InvoiceNo"
ws['B1'] = "StockCode"
ws['C1'] = "Description"
ws['D1'] = "Quantity"
ws['E1'] = "InvoiceDate"
ws['F1'] = "UnitPrice"
ws['G1'] = "CustomerID"
ws['H1'] = "Country"
ws.append([54365,85134,'whitehanging',6,11-11-2022,2.55,17850 ,'Unitedkingdom'])
import datetime
ws['A2'] = datetime.datetime.now()
wb.save("sam.csv")
